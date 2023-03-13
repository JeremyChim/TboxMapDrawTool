from datetime import datetime
from openpyxl import Workbook
import datetime

def write_to_excel(locations:list,
                   dtime:list):
    workbook = Workbook()
    worksheet_name = "打点数据"
    worksheet = workbook.create_sheet(title=worksheet_name,index=0)
    worksheet.column_dimensions['A'].width = 30 #设置表格宽度

    for i in range(1,len(locations)+1):
        worksheet.cell(i, 1).value = str(dtime[i-1])
        worksheet.cell(i, 2).value = str(locations[i-1])
        i += 1

    now_time = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    e_name = "locations_" + now_time + ".xlsx"
    workbook.save(e_name)
    workbook.close()

if __name__ == '__main__':
    _locations = [[22.614174538750753, 113.92599549079367], [22.61417520394314, 113.92599365710555]]
    print(type(_locations))
    _dtime = [datetime.datetime(2023, 2, 15, 11, 59, 59),datetime.datetime(2023, 2, 15, 12, 59, 59)]
    print(type(_dtime))
    write_to_excel(_locations, _dtime)